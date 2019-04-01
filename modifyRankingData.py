# import Workbook module
from openpyxl import Workbook
# import regex module
import re
# import stats module
import pandas as pd
import numpy as np
# import to debug to stop script, use: sys.exit(0)
import sys 

#sys.exit(0)


# load existing workbook ----
from openpyxl import load_workbook
# set file path
filepath="C:\Python27\myscripts\demo.xlsx"
# load demo.xlsx 
wb=load_workbook(filepath)
# select demo.xlsx
sheet=wb.active

# go through all rows of active sheet ----
# get max row count
max_row=sheet.max_row
# - print(max_row)

# Create Pattern pageType ----
pageType = re.compile(r'er.\w+\/(\w+)\/')

# iterate over all rows range(start,stop) ----
for i in range(2,max_row+1):
    # takes all column 3 rows --
    cell_obj=sheet.cell(row=i,column=3)
    # print URL cell value      
    # - print(cell_obj.value)
    # apply regex on URL string 
    mo = pageType.search(cell_obj.value)
    # extract pageType from URL group & write on column 4
    sheet.cell(row=i, column=4).value = mo.group(1)

# save workbook 
wb.save(filepath)
